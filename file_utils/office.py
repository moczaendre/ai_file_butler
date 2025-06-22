from docx import Document
from openpyxl import load_workbook
import os
from pathlib import Path
import shutil
import win32api
try:
    import win32com.client as win32
except ImportError:
    win32 = None

from file_utils.common import log, log_rename, is_file_locked  
from config import DEBUG, cfg

out_dir = cfg["office_output"]
out_dir.mkdir(parents=True, exist_ok=True)


def convert_doc_to_docx(input_path):
    """
    .doc => .docx konverzió
    """
    if is_file_locked(input_path):
        log(f"⚠️ A Word fájl zárolva:{input_path}", level="ERROR", module="office", to_console=True)
        return input_path
    if win32 is None:
        log("⚠️ A Word konverzióhoz szükséges a 'pywin32' csomag.", level="ERROR", module="office", to_console=True)
        return input_path

    input_path = Path(input_path).resolve()
    output_path = input_path.with_suffix('.docx')

    word = None
    doc = None

    try:
        if DEBUG["office"]:
            print(f"[DOC] Konvertálás elkezdve: {input_path}")
        word = win32.Dispatch("Word.Application")
        doc = word.Documents.Open(str(input_path))
        if DEBUG["office"]:
            print(f"[DOC] Megnyitva: {input_path}")
        doc.SaveAs(str(output_path), FileFormat=16)  # wdFormatDocumentDefault
        if DEBUG["office"]:
            print(f"[DOC] Mentve: {output_path}")
        
        log(f"[INFO] Átalakítva DOC → DOCX: {output_path.name}", module="office", to_console=True)
        return output_path
    except Exception as e:
        log(f"⚠️ DOC konverziós hiba: {e}", level="ERROR", module="office", to_console=True)
        return input_path  # visszatérünk az eredeti fájlra, hogy ne akadjon meg
    finally:
        if doc:
            doc.Close(SaveChanges=False)
        if word:
            word.Quit()

def convert_xls_to_xlsx(input_path):
    """
    .xls => .xlsx konverzió
    """
    if is_file_locked(input_path):
        log(f"⚠️ A Excel fájl zárolva:{input_path}", level="ERROR", module="office", to_console=True)
        return input_path
    if win32 is None:
        log("⚠️ A Word konverzióhoz szükséges a 'pywin32' csomag.", level="ERROR", module="office", to_console=True)
        return input_path

    input_path = Path(input_path).resolve()
    output_path = input_path.with_suffix('.xlsx')

    excel = None
    wb = None

    try:
        if DEBUG["office"]:
            print(f"[XLS] Konvertálás elkezdve: {input_path}")
        excel = win32.Dispatch("Excel.Application")
        wb = excel.Workbooks.Open(str(input_path))
        if DEBUG["office"]:
            print(f"[XLS] Megnyitva: {input_path}")
        wb.SaveAs(str(output_path), FileFormat=51)  # 51 = xlOpenXMLWorkbook (.xlsx)
        if DEBUG["office"]: 
            print(f"[XLS] Mentve: {output_path}")
        log(f"[INFO] Átalakítva XLS → XLSX: {output_path.name}", module="office", to_console=True)
        return output_path

    except Exception as e:
        log(f"⚠️ XLS konverziós hiba: {e}", level="ERROR", module="office", to_console=True)
        return input_path

    finally:
        if wb:
            wb.Close(SaveChanges=False)
        if excel:
            excel.Quit()



def read_docx(file_path):
    """
    DOCX fájl tartalmának szöveges kiolvasása.
    """
    try:
        doc = Document(file_path)
        full_text = "\n".join([para.text for para in doc.paragraphs])
        if DEBUG["office"]:
            log(f"[DOCX] Tartalom kivonat:\n{full_text[:500]}\n", level="DEBUG", module="office", to_console=True)
        return full_text
    except Exception as e:
        log(f"⚠️ Hiba DOCX fájlnál: {file_path.name} – {e}", level="ERROR", module="office", to_console=True)
        return ""

def read_xlsx(file_path):
    """
    XLSX fájl tartalmának szöveges kiolvasása.
    """
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        values = []
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            row_text = [str(cell) if cell is not None else "" for cell in row]
            values.append("\t".join(row_text))
        text = "\n".join(values)
        if DEBUG["office"]:
            log(f"[XLSX] Tartalom kivonat:\n{text[:500]}\n", module="office", level="DEBUG", to_console=True)
        return text
    except Exception as e:
        log(f"⚠️ Hiba XLSX fájlnál: {file_path.name} – {e}", level="ERROR", module="office", to_console=True)
        return ""

def move_file(file_path: Path):
    """
    Office fájlok áthelyezése
    """    
    target_path = out_dir / file_path.name
    shutil.move(str(file_path), str(target_path))
    log_rename(str(file_path), str(target_path))
    print(f"[OFFICE] Áthelyezve: {file_path.name} → {out_dir}/")

def process_office(file_path: Path):
    ext = file_path.suffix.lower()
    if ext == ".doc":
        converted = convert_doc_to_docx(str(file_path))
        read_docx(converted)
    elif ext == ".docx":
        read_docx(str(file_path))
    elif ext == ".xls":
        converted = convert_xls_to_xlsx(str(file_path))
        read_xlsx(converted)
    elif ext == ".xlsx":
        read_xlsx(str(file_path))
        
    move_file(file_path)